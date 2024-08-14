import openpyxl
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# Choose which file to run 0-3 (starting at 0)
file_number = 0
new_name = "_forPrint.xlsx"

# All files
xlsx_files = ["EYP_Ventilation_Calculator_v02 - CSH Admin and BoH.xlsx",
            "EYP_Ventilation_Calculator_v02 - CSH Civil Living Units.xlsx",
            "EYP_Ventilation_Calculator_v02 - CSH Hospital.xlsx",
            "EYP_Ventilation_Calculator_v02 - CSH Max Living Units.xlsx"]
# Set file name for script
xlsx = xlsx_files[file_number]

# Load the workbook for xlsx file name
wb = openpyxl.load_workbook(filename = xlsx)

# Initialize formatting styles for no_fill and no_border
no_fill = openpyxl.styles.PatternFill(fill_type=None)
orange_fill = openpyxl.styles.PatternFill("solid", fgColor="FFCC99")
side = openpyxl.styles.Side(border_style=None)
no_border = openpyxl.styles.borders.Border(
    left=side, 
    right=side, 
    top=side, 
    bottom=side,
)

# Iterate through each sheet of the workbook
for sheet in wb.worksheets:
    # Change print options for all sheets
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToHeight = False # Fit page to columns instead of height

    if "AHU" not in sheet.title:
        sheet.sheet_state = "hidden"
    elif "Multizone AHU BLANK TEMPLATE" in sheet.title:
        sheet.sheet_state = "hidden"
        
    if "AHU" in sheet.title: # Work on only sheets that have sheet name with "AHU" in it
        # Copy internal info to hidden region for printing
        if sheet['I2'].value is not None and sheet['S2'].value is None:
            # Copy values from the cells into hidden columns
            sheet['S2'].value = sheet['I2'].value
            sheet['S4'].value = sheet['I4'].value
            sheet['S5'].value = sheet['I5'].value
            sheet['S9'].value = sheet['I9'].value
            sheet['S10'].value = sheet['I10'].value
            sheet['T10'].value = sheet['J10'].value
            sheet['U2'].value = sheet['K2'].value
            sheet['U4'].value = sheet['K4'].value
            sheet['U5'].value = sheet['K5'].value
            sheet['V2'].value = sheet['L2'].value
            sheet['Z2'].value = sheet['P2'].value
            sheet['Z3'].value = sheet['P3'].value
            sheet['Z4'].value = sheet['P4'].value
            sheet['Z5'].value = sheet['P5'].value
            sheet['Z6'].value = sheet['P6'].value
            sheet['AA2'].value = sheet['Q2'].value
            sheet['AA3'].value = sheet['Q3'].value
            sheet['AA4'].value = sheet['Q4'].value
            sheet['AA5'].value = sheet['Q5'].value
            sheet['AA6'].value = sheet['Q6'].value
            sheet.unmerge_cells('I2:J3')
            sheet.unmerge_cells('I4:J4')
            sheet.unmerge_cells('I5:J5')
            sheet.unmerge_cells('K2:K3')
            sheet.unmerge_cells('K4:N4')
            sheet.unmerge_cells('L2:N3')
            # Remove values from existing cells
            sheet['I2'].value = None
            sheet['I4'].value = None
            sheet['I5'].value = None
            sheet['I9'].value = None
            sheet['I10'].value = None
            sheet['J10'].value = None
            sheet['K2'].value = None
            sheet['K4'].value = None
            sheet['K5'].value = None
            sheet['L2'].value = None
            sheet['P2'].value = None
            sheet['P3'].value = None
            sheet['P4'].value = None
            sheet['P5'].value = None
            sheet['P6'].value = None
            sheet['Q2'].value = None
            sheet['Q3'].value = None
            sheet['Q4'].value = None
            sheet['Q5'].value = None
            sheet['Q6'].value = None
            # Remove formating for cells
            sheet['I2'].fill = no_fill
            sheet['I4'].fill = no_fill
            sheet['I5'].fill = no_fill
            sheet['I9'].fill = no_fill
            sheet['I10'].fill = no_fill
            sheet['J10'].fill = no_fill
            sheet['K2'].fill = no_fill
            sheet['K4'].fill = no_fill
            sheet['K5'].fill = no_fill
            sheet['L2'].fill = no_fill
            sheet['P2'].fill = no_fill
            sheet['P3'].fill = no_fill
            sheet['P4'].fill = no_fill
            sheet['P5'].fill = no_fill
            sheet['P6'].fill = no_fill
            sheet['Q2'].fill = no_fill
            sheet['Q3'].fill = no_fill
            sheet['Q4'].fill = no_fill
            sheet['Q5'].fill = no_fill
            sheet['Q6'].fill = no_fill
            sheet['I2'].border = no_border
            sheet['I4'].border = no_border
            sheet['I5'].border = no_border
            sheet['I9'].border = no_border
            sheet['I10'].border = no_border
            sheet['J3'].border = no_border
            sheet['J4'].border = no_border
            sheet['J5'].border = no_border
            sheet['J10'].border = no_border
            sheet['K2'].border = no_border
            sheet['K3'].border = no_border
            sheet['K4'].border = no_border
            sheet['K5'].border = no_border
            sheet['L2'].border = no_border
            sheet['L3'].border = no_border
            sheet['P2'].border = no_border
            sheet['P3'].border = no_border
            sheet['P4'].border = no_border
            sheet['P5'].border = no_border
            sheet['P6'].border = no_border
            sheet['Q2'].border = no_border
            sheet['Q3'].border = no_border
            sheet['Q4'].border = no_border
            sheet['Q5'].border = no_border
            sheet['Q6'].border = no_border
            # Correct formulas after moving
            sheet['C9'] = "=S10"
            sheet['C12'] = "=IF(U2=\"Appendix A\",U5,IF(U2=\"Table 6-3\",VLOOKUP(C11,'Ev System Ventilation Effic'!$A$2:$B$7,2,TRUE)))"
            sheet['U4'] = "=IF(U2=\"Table 6-3\",VLOOKUP(MAX(S18:S1048576),S18:Z1048576,8,FALSE),IF(U2=\"Appendix A\",VLOOKUP(MIN(Y18:Y1048576),Y18:Z1048576,2,FALSE),\"\"))"
            sheet['U5'] = "=VLOOKUP(U4,A18:Z1048576,25,FALSE)"

        # Hide columns not for printing
        sheet.column_dimensions.group(start='S', end='AF', hidden=True)
        
        # Iterate through rows to determine those that are to be hidden
        rows_with_content = 0
        for row in sheet.iter_rows(min_row=18,
                                    max_row=305,
                                    min_col=1,
                                    max_col=6,
                                    values_only=True,):
            if row[0] is not None:
                # print(row[0])
                rows_with_content = rows_with_content + 1
                # row[2].fill = orange_fill
        
        # print(rows_with_content)
        # Hide rows with None values determined from row_with_content count
        sheet.row_dimensions.group(start=18+rows_with_content, end=311, hidden=True)


wb.save(xlsx.replace(".xlsx", new_name))
print("Saved " + xlsx.replace(".xlsx", new_name))